from xlrd import open_workbook
import numpy as np
import pandas as pd
from colormap import rgb2hex
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill
def color_coordinated(workbook, sheet_name):
    wb = open_workbook(workbook, formatting_info=True)
    sheet = wb.sheet_by_name(sheet_name)
    color_coordinates = []
    for row in range(sheet.nrows):
        for column in range(sheet.ncols):
            xfx = sheet.cell_xf_index(row, column)
            xf = wb.xf_list[xfx]
            bgx = xf.background.pattern_colour_index
            pattern_colour = wb.colour_map[bgx]
            if pattern_colour is not None:
                print((row, column), pattern_colour)
                color_coordinate = {
                    'row': row, 
                    'column': column,
                    'rbg': rgb2hex(pattern_colour[0], pattern_colour[1], pattern_colour[2])}
                color_coordinates.append(color_coordinate)
    return color_coordinates



def fill_cell(workbook, sheet_name, color_coordinate):
    ws = workbook.sheets[sheet_name]
    fill_color = PatternFill(start_color=color_coordinate.rbg,
                    end_color=color_coordinate.rbg,
                    fill_type='solid')
    ws[f'${get_column_letter(color_coordinate.column - 1)}'] = fill_cell.fill = fill_color